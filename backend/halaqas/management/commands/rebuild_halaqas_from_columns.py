from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import Profile
from halaqas.models import Category, Halaqa, HalaqaMembership, Teacher
from students.models import Student


ARABIC_DIGIT_TRANSLATION = str.maketrans(
    "\u0660\u0661\u0662\u0663\u0664\u0665\u0666\u0667\u0668\u0669\u06f0\u06f1\u06f2\u06f3\u06f4\u06f5\u06f6\u06f7\u06f8\u06f9",
    "01234567890123456789",
)
DIACRITICS_RE = re.compile(r"[\u0610-\u061a\u064b-\u065f\u0670\u06d6-\u06ed]")
SPACE_RE = re.compile(r"\s+")

HALAQA_PREFIX = "\u062d\u0644\u0642\u0629 "
BAKALORIA_WORDS = {
    "\u0628\u0643\u0627\u0644\u0648\u0631\u064a\u0627",
    "\u0628\u0643\u0644\u0648\u0631\u064a\u0627",
    "\u0627\u0644\u0628\u0643\u0627\u0644\u0648\u0631\u064a\u0627",
    "\u0627\u0644\u0628\u0643\u0644\u0648\u0631\u064a\u0627",
}

CATEGORY_SPECS = {
    "1": {
        "name": "\u0627\u0644\u0641\u0626\u0629 \u0627\u0644\u0623\u0648\u0644\u0649",
        "grade_span": "\u0627\u0644\u0635\u0641\u0648\u0641 2\u060c 3\u060c 4",
        "display_order": 1,
        "grades": {2, 3, 4},
        "create_code": "C1",
    },
    "2": {
        "name": "\u0627\u0644\u0641\u0626\u0629 \u0627\u0644\u062b\u0627\u0646\u064a\u0629",
        "grade_span": "\u0627\u0644\u0635\u0641\u0648\u0641 5\u060c 6",
        "display_order": 2,
        "grades": {5, 6},
        "create_code": "C2",
    },
    "3": {
        "name": "\u0627\u0644\u0641\u0626\u0629 \u0627\u0644\u062b\u0627\u0644\u062b\u0629",
        "grade_span": "\u0627\u0644\u0635\u0641\u0648\u0641 7\u060c 8",
        "display_order": 3,
        "grades": {7, 8},
        "create_code": "C3",
    },
    "4": {
        "name": "\u0627\u0644\u0641\u0626\u0629 \u0627\u0644\u0631\u0627\u0628\u0639\u0629",
        "grade_span": "\u0627\u0644\u0635\u0641 9",
        "display_order": 4,
        "grades": {9},
        "create_code": "C4",
    },
    "5": {
        "name": "\u0627\u0644\u0641\u0626\u0629 \u0627\u0644\u062e\u0627\u0645\u0633\u0629",
        "grade_span": "\u0627\u0644\u0635\u0641\u0648\u0641 10\u060c 11\u060c 12 / \u0627\u0644\u0628\u0643\u0627\u0644\u0648\u0631\u064a\u0627",
        "display_order": 5,
        "grades": {10, 11, 12},
        "create_code": "C5",
    },
}

GRADE_TO_CATEGORY = {
    grade: code
    for code, spec in CATEGORY_SPECS.items()
    for grade in spec["grades"]
}


@dataclass
class ColumnHalaqa:
    column_index: int
    column_letter: str
    teacher_names: list[str]
    student_names: list[str]
    grade_raw: str
    grades: list[int]
    category_code: str | None
    halaqa_name: str


@dataclass
class Issue:
    kind: str
    column: str
    value: str
    message: str


class VirtualRecord:
    def __init__(self, name, pk=None, id=None):
        self.name = name
        self.full_name = name
        self.pk = pk
        self.id = id


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFKC", str(value)).translate(ARABIC_DIGIT_TRANSLATION)
    text = text.replace("\u200f", " ").replace("\u200e", " ").replace("\u0640", "")
    return SPACE_RE.sub(" ", text).strip()


def normalize_for_match(value) -> str:
    text = clean_text(value)
    text = DIACRITICS_RE.sub("", text)
    text = re.sub("[\u0625\u0623\u0622\u0627]", "\u0627", text)
    text = text.replace("\u0649", "\u064a").replace("\u0629", "\u0647")
    return SPACE_RE.sub(" ", text).strip().lower()


def split_teacher_names(value) -> list[str]:
    raw = "" if value is None else str(value)
    names = []
    for part in re.split(r"[+\uff0b\r\n]+", raw):
        name = clean_text(part)
        if name:
            names.append(name)
    return unique_by_normalized(names)


def unique_by_normalized(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        key = normalize_for_match(value)
        if key and key not in seen:
            seen.add(key)
            result.append(value)
    return result


def parse_grade(value) -> list[int]:
    normalized = normalize_for_match(value)
    if not normalized:
        return []
    grades = {int(match) for match in re.findall(r"\d+", normalized) if 2 <= int(match) <= 12}
    if any(word in normalized for word in BAKALORIA_WORDS):
        grades.add(12)
    return sorted(grades)


def category_for_grades(grades: list[int]) -> str | None:
    if not grades:
        return None
    codes = {GRADE_TO_CATEGORY.get(grade) for grade in grades}
    codes.discard(None)
    if len(codes) == 1:
        return codes.pop()
    return None


def category_name(category_code: str | None) -> str:
    if not category_code:
        return "UNKNOWN"
    return CATEGORY_SPECS[category_code]["name"]


class Rebuilder:
    def __init__(self, command: BaseCommand, options):
        self.command = command
        self.stdout = command.stdout
        self.style = command.style
        self.options = options
        self.dry_run = options["dry_run"]
        self.issues: list[Issue] = []
        self.stats = Counter()
        self.teacher_index = self._build_index(Teacher.objects.all(), "full_name")
        self.student_index = self._build_index(Student.objects.all(), "name")
        self.planned_teachers = {}
        self.planned_students = {}
        self.planned_usernames = set()
        self.category_cache = {}
        self.default_birth_date = date(1900, 1, 1)

    def _build_index(self, queryset, attr: str):
        index = defaultdict(list)
        for obj in queryset:
            index[normalize_for_match(getattr(obj, attr))].append(obj)
        return index

    def run(self, columns: list[ColumnHalaqa], sheet_name: str):
        duplicate_students = self._duplicate_names(columns, "student")
        unknown_grade_columns = [column for column in columns if not column.category_code]

        for value, places in duplicate_students.items():
            self._add_issue("duplicate_student_in_sheet", ", ".join(places), value, "Student appears in more than one column.")
        for column in unknown_grade_columns:
            self._add_issue("unknown_grade", column.column_letter, column.grade_raw, "Grade does not map to a known category.")

        self._print_parsed_summary(columns, sheet_name)
        self._preview_database_changes(columns)

        blocker_kinds = {
            "duplicate_student_in_sheet",
            "unknown_grade",
            "missing_category",
            "missing_teacher",
            "missing_student",
            "duplicate_teacher_in_db",
            "duplicate_student_in_db",
            "duplicate_student_exact_in_db",
        }
        blockers = [issue for issue in self.issues if issue.kind in blocker_kinds]
        if self.options["conflicts_csv"]:
            self._write_conflicts_csv(self.options["conflicts_csv"])

        if self.dry_run:
            self._print_summary()
            self.stdout.write(self.style.WARNING("Dry run only. No database changes were written."))
            return

        if blockers:
            self._print_summary()
            raise CommandError("Commit blocked because duplicates or unknown grades were found. Resolve conflicts and rerun.")

        self.stats = Counter()
        self.category_cache = {}
        self.planned_teachers = {}
        self.planned_students = {}
        with transaction.atomic():
            self._commit(columns)

        self._print_summary()
        self.stdout.write(self.style.SUCCESS("Commit complete. Active halaqa relationships were rebuilt without deleting history."))

    def _duplicate_names(self, columns: list[ColumnHalaqa], name_type: str):
        locations = defaultdict(list)
        display_values = {}
        for column in columns:
            values = column.student_names if name_type == "student" else column.teacher_names
            for value in values:
                key = normalize_for_match(value)
                if not key:
                    continue
                locations[key].append(column.column_letter)
                display_values.setdefault(key, value)
        return {
            display_values[key]: places
            for key, places in locations.items()
            if len(set(places)) > 1
        }

    def _preview_database_changes(self, columns: list[ColumnHalaqa]):
        desired_student_targets = {}
        for column in columns:
            if not column.category_code:
                continue
            self._resolve_category(column.category_code, preview=True)
            self._resolve_halaqa(column, preview=True)
            for teacher_name in column.teacher_names:
                self._resolve_teacher(teacher_name, column.column_letter, preview=True)
            for student_name in column.student_names:
                student = self._resolve_student(student_name, column, preview=True)
                if student and not isinstance(student, VirtualRecord):
                    current = (
                        HalaqaMembership.objects.filter(student=student, is_active=True)
                        .select_related("halaqa")
                        .order_by("-join_date", "-id")
                        .first()
                    )
                    if current and current.halaqa.name != column.halaqa_name:
                        desired_student_targets[student.name] = f"{current.halaqa.name} -> {column.halaqa_name}"
        self.stats["student_moves"] = len(desired_student_targets)
        for name, move in desired_student_targets.items():
            self._add_issue("student_move", "", name, move)

    def _commit(self, columns: list[ColumnHalaqa]):
        imported_halaqas = []
        desired_students_by_halaqa = defaultdict(set)

        for column in columns:
            category = self._resolve_category(column.category_code, preview=False)
            halaqa = self._resolve_halaqa(column, preview=False, category=category)
            imported_halaqas.append(halaqa)

            teachers = [
                teacher
                for teacher_name in column.teacher_names
                if (teacher := self._resolve_teacher(teacher_name, column.column_letter, preview=False))
            ]
            self._set_halaqa_teachers_without_assignment_sync(halaqa, teachers)

            for student_name in column.student_names:
                student = self._resolve_student(student_name, column, preview=False, category=category, halaqa=halaqa)
                if not student:
                    continue
                desired_students_by_halaqa[halaqa.id].add(student.id)
                self._activate_membership(student, halaqa, category, column.grade_raw)

        today = timezone.localdate()
        for halaqa in imported_halaqas:
            desired_ids = desired_students_by_halaqa[halaqa.id]
            old_memberships = HalaqaMembership.objects.filter(
                halaqa=halaqa,
                is_active=True,
            ).exclude(student_id__in=desired_ids)
            changed = old_memberships.filter(end_date__isnull=True).update(end_date=today)
            changed += old_memberships.update(is_active=False)
            if changed:
                self.stats["student_memberships_deactivated"] += changed

        Student.objects.filter(
            halaqa__in=imported_halaqas,
        ).exclude(id__in={student_id for ids in desired_students_by_halaqa.values() for student_id in ids}).update(
            halaqa=None
        )

    def _set_halaqa_teachers_without_assignment_sync(self, halaqa: Halaqa, teachers: list[Teacher]):
        through_model = Halaqa.teachers.through
        desired_teacher_ids = {teacher.pk for teacher in teachers}
        through_model.objects.filter(halaqa_id=halaqa.pk).exclude(teacher_id__in=desired_teacher_ids).delete()
        for teacher_id in desired_teacher_ids:
            _, created = through_model.objects.get_or_create(teacher_id=teacher_id, halaqa_id=halaqa.pk)
            if created:
                self.stats["teacher_halaqa_links_created"] += 1
            else:
                self.stats["teacher_halaqa_links_reuse"] += 1
        self.stats["teacher_halaqa_sets"] += 1

    def _resolve_category(self, code: str | None, preview: bool):
        if not code:
            return None
        if code in self.category_cache:
            return self.category_cache[code]

        spec = CATEGORY_SPECS[code]
        category = Category.objects.filter(name=spec["name"]).first()
        if category:
            self.stats["categories_reuse"] += 1
            self.category_cache[code] = category
            return category

        if not self.options["create_missing_categories"]:
            self.stats["categories_missing"] += 1
            self._add_issue("missing_category", "", spec["name"], "Category is missing. Add --create-missing-categories to create it.")
            return None

        self.stats["categories_create"] += 1
        if preview:
            category = VirtualRecord(spec["name"])
        else:
            category = Category.objects.create(
                code=self._available_category_code(spec["create_code"]),
                name=spec["name"],
                grade_span=spec["grade_span"],
                display_order=spec["display_order"],
                is_special=False,
                notes="",
            )
        self.category_cache[code] = category
        return category

    def _available_category_code(self, preferred: str) -> str:
        if not Category.objects.filter(code=preferred).exists():
            return preferred
        for first in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            for second in "12345":
                code = f"{first}{second}"
                if not Category.objects.filter(code=code).exists():
                    return code
        raise CommandError("No available two-character category code.")

    def _resolve_halaqa(self, column: ColumnHalaqa, preview: bool, category=None):
        halaqa = Halaqa.objects.filter(name=column.halaqa_name).first()
        if halaqa:
            self.stats["halaqas_reuse"] += 1
            if category and not isinstance(category, VirtualRecord) and halaqa.category_id != category.id:
                self.stats["halaqas_update_category"] += 1
                if not preview:
                    halaqa.category = category
                    halaqa.is_active = True
                    halaqa.save(update_fields=["category", "is_active"])
            return halaqa

        self.stats["halaqas_create"] += 1
        if preview:
            return VirtualRecord(column.halaqa_name)
        return Halaqa.objects.create(name=column.halaqa_name[:100], category=category, is_active=True)

    def _resolve_teacher(self, teacher_name: str, column: str, preview: bool):
        normalized = normalize_for_match(teacher_name)
        matches = self.teacher_index.get(normalized, [])
        if len(matches) == 1:
            self.stats["teachers_reuse"] += 1
            return matches[0]
        if len(matches) > 1:
            self.stats["teacher_name_conflicts"] += 1
            self._add_issue("duplicate_teacher_in_db", column, teacher_name, "More than one teacher has this normalized name.")
            return None
        if normalized in self.planned_teachers:
            self.stats["teachers_reuse_planned"] += 1
            return self.planned_teachers[normalized]
        if not self.options["create_missing_teachers"]:
            self.stats["teachers_missing"] += 1
            self._add_issue("missing_teacher", column, teacher_name, "Teacher is missing. Add --create-missing-teachers to create it.")
            return None

        self.stats["teachers_create"] += 1
        if preview:
            teacher = VirtualRecord(teacher_name)
        else:
            User = get_user_model()
            user = User(username=self._next_username("teacher"), first_name=teacher_name[:150], email="")
            user.set_unusable_password()
            user.save()
            profile, _ = Profile.objects.get_or_create(user=user)
            if profile.role != "teacher":
                profile.role = "teacher"
                profile.save(update_fields=["role"])
            teacher = Teacher.objects.create(user=user, full_name=teacher_name, phone="", qualification="")
            self.teacher_index[normalized].append(teacher)
        self.planned_teachers[normalized] = teacher
        return teacher

    def _resolve_student(self, student_name: str, column: ColumnHalaqa, preview: bool, category=None, halaqa=None):
        normalized = normalize_for_match(student_name)
        matches = self.student_index.get(normalized, [])
        if len(matches) == 1:
            self.stats["students_reuse"] += 1
            return matches[0]
        if len(matches) > 1:
            self.stats["student_name_conflicts"] += 1
            exact_matches = [student for student in matches if clean_text(student.name) == clean_text(student_name)]
            issue_kind = "duplicate_student_exact_in_db" if len(exact_matches) > 1 else "duplicate_student_in_db"
            details = "; ".join(self._student_match_detail(student) for student in matches)
            self._add_issue(
                issue_kind,
                column.column_letter,
                student_name,
                f"More than one existing student matches this name. Matches: {details}. Manual resolution required.",
            )
            if not exact_matches and self.options["allow_duplicate_student_name_create"]:
                return self._create_student(student_name, column, preview, category=category, halaqa=halaqa)
            return None
        if normalized in self.planned_students:
            self.stats["students_reuse_planned"] += 1
            return self.planned_students[normalized]
        if not self.options["create_missing_students"]:
            self.stats["students_missing"] += 1
            self._add_issue("missing_student", column.column_letter, student_name, "Student is missing. Add --create-missing-students to create it.")
            return None

        return self._create_student(student_name, column, preview, category=category, halaqa=halaqa)

    def _student_match_detail(self, student: Student) -> str:
        current_membership = (
            HalaqaMembership.objects.filter(student=student, is_active=True)
            .select_related("halaqa")
            .order_by("-join_date", "-id")
            .first()
        )
        current_halaqa = current_membership.halaqa if current_membership else student.halaqa
        current_halaqa_name = current_halaqa.name if current_halaqa else "no current halaqa"
        return f"id={student.pk}, name={student.name}, current_halaqa={current_halaqa_name}"

    def _create_student(self, student_name: str, column: ColumnHalaqa, preview: bool, category=None, halaqa=None):
        normalized = normalize_for_match(student_name)
        self.stats["students_create"] += 1
        if preview:
            student = VirtualRecord(student_name)
        else:
            student = Student.objects.create(
                name=student_name,
                birth_date=self.default_birth_date,
                parent=None,
                parent_phone="",
                address="",
                grade=column.grade_raw,
                category=category,
                halaqa=halaqa,
                enrollment_date=timezone.localdate(),
                previous_memorization_amount=0,
            )
            self.student_index[normalized].append(student)
        self.planned_students[normalized] = student
        return student

    def _activate_membership(self, student: Student, halaqa: Halaqa, category: Category, grade: str):
        changed_fields = []
        if student.grade != grade:
            student.grade = grade
            changed_fields.append("grade")
        if student.category_id != category.id:
            student.category = category
            changed_fields.append("category")
        if student.halaqa_id != halaqa.id:
            student.halaqa = halaqa
            changed_fields.append("halaqa")
        if changed_fields:
            student.save(update_fields=changed_fields)
            self.stats["students_update"] += 1

        membership = HalaqaMembership.objects.filter(student=student, halaqa=halaqa).first()
        if membership:
            if not membership.is_active or membership.end_date:
                membership.is_active = True
                membership.end_date = None
                membership.save(update_fields=["is_active", "end_date"])
                self.stats["student_memberships_reactivated"] += 1
            else:
                self.stats["student_memberships_reuse"] += 1
            return

        HalaqaMembership.objects.create(student=student, halaqa=halaqa, is_active=True)
        self.stats["student_memberships_created"] += 1

    def _next_username(self, prefix: str) -> str:
        User = get_user_model()
        counter = 1
        while True:
            username = f"{prefix}_{counter:03d}"
            if username not in self.planned_usernames and not User.objects.filter(username=username).exists():
                self.planned_usernames.add(username)
                return username
            counter += 1

    def _add_issue(self, kind: str, column: str, value: str, message: str):
        issue = Issue(kind, column, clean_text(value), message)
        key = (issue.kind, issue.column, normalize_for_match(issue.value), issue.message)
        if not hasattr(self, "_issue_keys"):
            self._issue_keys = set()
        if key in self._issue_keys:
            return
        self._issue_keys.add(key)
        self.issues.append(issue)

    def _print_parsed_summary(self, columns: list[ColumnHalaqa], sheet_name: str):
        self.stdout.write(self.style.NOTICE(f"Sheet used: {sheet_name}"))
        self.stdout.write(self.style.NOTICE(f"Parsed halaqas/columns: {len(columns)}"))
        self.stdout.write("")
        self.stdout.write("Column  Grade       Category        Students  Halaqa / teachers")
        self.stdout.write("-" * 78)
        for column in columns:
            teachers = " + ".join(column.teacher_names)
            self.stdout.write(
                f"{column.column_letter:<7} {column.grade_raw:<11} {category_name(column.category_code):<15} "
                f"{len(column.student_names):>8}  {column.halaqa_name} ({teachers})"
            )

    def _print_summary(self):
        self.stdout.write("")
        self.stdout.write(self.style.NOTICE("Import summary"))
        for key in [
            "categories_reuse",
            "categories_create",
            "categories_missing",
            "halaqas_reuse",
            "halaqas_create",
            "halaqas_update_category",
            "teachers_reuse",
            "teachers_create",
            "teachers_missing",
            "teacher_name_conflicts",
            "teacher_halaqa_sets",
            "teacher_halaqa_links_created",
            "teacher_halaqa_links_reuse",
            "students_reuse",
            "students_create",
            "students_missing",
            "student_name_conflicts",
            "student_moves",
            "students_update",
            "student_memberships_created",
            "student_memberships_reactivated",
            "student_memberships_reuse",
            "student_memberships_deactivated",
        ]:
            self.stdout.write(f"- {key}: {self.stats.get(key, 0)}")

        if self.issues:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING(f"Conflicts/warnings: {len(self.issues)}"))
            for issue in self.issues[:80]:
                column = f"{issue.column}: " if issue.column else ""
                self.stdout.write(f"- [{issue.kind}] {column}{issue.value} - {issue.message}")
            if len(self.issues) > 80:
                self.stdout.write(f"... {len(self.issues) - 80} more issues. Use --conflicts-csv for the full list.")

    def _write_conflicts_csv(self, path_value: str):
        path = Path(path_value)
        if not path.is_absolute():
            path = Path.cwd() / path
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["kind", "column", "value", "message"])
            writer.writeheader()
            for issue in self.issues:
                writer.writerow(
                    {
                        "kind": issue.kind,
                        "column": issue.column,
                        "value": issue.value,
                        "message": issue.message,
                    }
                )
        self.stdout.write(self.style.SUCCESS(f"Conflicts CSV written: {path}"))


class Command(BaseCommand):
    help = "Rebuild active halaqa/teacher/student/category assignments from an Excel sheet where each column is one halaqa."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the Excel workbook.")
        parser.add_argument("--sheet", help="Sheet name. If omitted, prefer 'j', then 'Sheet2', then a matching sheet.")
        mode = parser.add_mutually_exclusive_group()
        mode.add_argument("--dry-run", action="store_true", help="Parse and report without writing database changes.")
        mode.add_argument("--commit", action="store_true", help="Write assignment changes to the database.")
        parser.add_argument("--create-missing-teachers", action="store_true")
        parser.add_argument("--create-missing-students", action="store_true")
        parser.add_argument("--create-missing-categories", action="store_true")
        parser.add_argument(
            "--allow-duplicate-student-name-create",
            action="store_true",
            help=(
                "Create a new student when only normalized-name matches are duplicated. "
                "Exact duplicate existing names still block commit."
            ),
        )
        parser.add_argument("--conflicts-csv", default="", help="Optional path for conflicts/warnings CSV.")

    def handle(self, *args, **options):
        options["dry_run"] = not options["commit"]
        workbook_path = Path(options["file"])
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path, read_only=False, data_only=True)
        try:
            sheet_name = self._select_sheet(workbook, options.get("sheet"))
            columns = self._parse_sheet(workbook[sheet_name])
            if not columns:
                raise CommandError("No non-empty halaqa columns were parsed.")
        finally:
            workbook.close()

        rebuilder = Rebuilder(self, options)
        rebuilder.run(columns, sheet_name)

    def _select_sheet(self, workbook, requested_sheet: str | None) -> str:
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                raise CommandError(f"Sheet not found: {requested_sheet}. Available: {', '.join(workbook.sheetnames)}")
            return requested_sheet
        for preferred in ("j", "Sheet2"):
            if preferred in workbook.sheetnames:
                return preferred
        for sheet_name in workbook.sheetnames:
            if self._parse_sheet(workbook[sheet_name]):
                return sheet_name
        raise CommandError("Could not find a sheet matching the expected column format.")

    def _parse_sheet(self, worksheet) -> list[ColumnHalaqa]:
        columns = []
        for column_cells in worksheet.iter_cols():
            non_empty = []
            for cell in column_cells:
                value = clean_text(cell.value)
                if value:
                    non_empty.append((cell.row, value))
            if len(non_empty) < 2:
                continue

            teacher_raw = non_empty[0][1]
            grade_raw = non_empty[-1][1]
            students = unique_by_normalized([value for _, value in non_empty[1:-1]])
            teachers = split_teacher_names(teacher_raw)
            grades = parse_grade(grade_raw)
            category_code = category_for_grades(grades)
            halaqa_name = (HALAQA_PREFIX + " + ".join(teachers))[:100] if teachers else f"Column {column_cells[0].column_letter}"
            columns.append(
                ColumnHalaqa(
                    column_index=column_cells[0].column,
                    column_letter=column_cells[0].column_letter,
                    teacher_names=teachers,
                    student_names=students,
                    grade_raw=grade_raw,
                    grades=grades,
                    category_code=category_code,
                    halaqa_name=halaqa_name,
                )
            )
        return columns
