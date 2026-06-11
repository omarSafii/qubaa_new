from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from accounts.models import Profile
from halaqas.models import Category, Halaqa, HalaqaMembership, Teacher, TeacherAssignment
from students.models import Student


ARABIC_DIGIT_TRANSLATION = str.maketrans(
    "٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹",
    "01234567890123456789",
)
DIACRITICS_RE = re.compile(r"[\u0610-\u061a\u064b-\u065f\u0670\u06d6-\u06ed]")
SPACE_RE = re.compile(r"\s+")

CATEGORY_SPECS = {
    "1": {
        "name": "الفئة الأولى",
        "grade_span": "الصفوف 2، 3، 4",
        "display_order": 1,
        "is_special": False,
        "grades": {2, 3, 4},
    },
    "2": {
        "name": "الفئة الثانية",
        "grade_span": "الصفوف 5، 6",
        "display_order": 2,
        "is_special": False,
        "grades": {5, 6},
    },
    "3": {
        "name": "الفئة الثالثة",
        "grade_span": "الصفوف 7، 8",
        "display_order": 3,
        "is_special": False,
        "grades": {7, 8},
    },
    "4": {
        "name": "الفئة الرابعة",
        "grade_span": "الصف 9",
        "display_order": 4,
        "is_special": False,
        "grades": {9},
    },
    "5": {
        "name": "الفئة الخامسة",
        "grade_span": "الصفوف 10، 11، 12 / البكالوريا",
        "display_order": 5,
        "is_special": False,
        "grades": {10, 11, 12},
    },
}

GRADE_TO_CATEGORY = {
    grade: code
    for code, spec in CATEGORY_SPECS.items()
    for grade in spec["grades"]
}

GRADE_WORDS = {
    "الثاني": 2,
    "ثاني": 2,
    "الثانيه": 2,
    "الثالث": 3,
    "ثالث": 3,
    "الثالثه": 3,
    "الرابع": 4,
    "رابع": 4,
    "الرابعه": 4,
    "الخامس": 5,
    "خامس": 5,
    "الخامسه": 5,
    "السادس": 6,
    "سادس": 6,
    "السادسه": 6,
    "السابع": 7,
    "سابع": 7,
    "السابعه": 7,
    "الثامن": 8,
    "ثامن": 8,
    "الثامنه": 8,
    "التاسع": 9,
    "تاسع": 9,
    "التاسعه": 9,
    "العاشر": 10,
    "عاشر": 10,
    "العاشره": 10,
    "الحادي عشر": 11,
    "حادي عشر": 11,
    "احد عشر": 11,
    "الحاديه عشر": 11,
    "البكالوريا": 12,
    "بكالوريا": 12,
    "بكلوريا": 12,
    "الثاني عشر": 12,
    "ثاني عشر": 12,
}

TEACHER_KEYWORDS = (
    "استاذ",
    "أستاذ",
    "الأستاذ",
    "الاستاذ",
    "معلم",
    "معلمة",
    "شيخ",
    "شيخة",
    "مدرس",
    "مدرسة",
)
HALAQA_KEYWORDS = ("حلقة", "الحلقة", "حلقه", "الحلقه")
GENERIC_WORDS = {
    "م",
    "رقم",
    "الرقم",
    "اسم",
    "الاسم",
    "اسم الطالب",
    "اسم الطالبة",
    "الطالب",
    "الطالبة",
    "الطلاب",
    "الطالبات",
    "الصف",
    "صف",
    "الفئة",
    "الفئه",
    "الحلقة",
    "حلقه",
    "المعلم",
    "المعلمة",
    "الأستاذ",
    "الاستاذ",
    "ملاحظات",
}


@dataclass
class Issue:
    kind: str
    row: int | None
    value: str
    message: str


@dataclass
class BlockStart:
    row_index: int
    anchor_col: int
    grade_col: int | None
    grades: list[int]
    category_code: str | None
    category_source: str


@dataclass
class ImportBlock:
    row_number: int
    left_col: int
    right_col: int
    start_key: tuple[int, int, int | None]
    grades: list[int]
    grade_label: str
    category_code: str | None
    teacher_names: list[str]
    halaqa_name: str
    students: list[str]
    raw_header: str
    issues: list[Issue] = field(default_factory=list)


@dataclass
class VirtualRecord:
    name: str
    pk: int | None = None
    id: int | None = None
    category_id: int | None = None
    full_name: str = ""


class ImportRollback(Exception):
    pass


def normalize_for_match(value) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).translate(ARABIC_DIGIT_TRANSLATION)
    text = DIACRITICS_RE.sub("", text)
    text = text.replace("ـ", "")
    text = re.sub("[إأآا]", "ا", text)
    text = text.replace("ى", "ي").replace("ة", "ه")
    text = SPACE_RE.sub(" ", text)
    return text.strip().lower()


def clean_display_text(value) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).translate(ARABIC_DIGIT_TRANSLATION)
    text = text.replace("\u200f", " ").replace("\u200e", " ")
    text = text.replace("ـ", "")
    text = SPACE_RE.sub(" ", text)
    return text.strip()


def cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (date,)):
        return value.isoformat()
    return clean_display_text(value)


def is_generic_text(text: str) -> bool:
    normalized = normalize_for_match(text)
    return not normalized or normalized in {normalize_for_match(word) for word in GENERIC_WORDS}


def contains_teacher_keyword(text: str) -> bool:
    normalized = normalize_for_match(text)
    return any(normalize_for_match(keyword) in normalized for keyword in TEACHER_KEYWORDS)


def contains_halaqa_keyword(text: str) -> bool:
    normalized = normalize_for_match(text)
    return any(normalize_for_match(keyword) in normalized for keyword in HALAQA_KEYWORDS)


def split_cell_parts(text: str) -> list[str]:
    text = clean_display_text(text)
    text = text.replace("\r", "\n")
    parts = []
    for line in re.split(r"[\n؛;]+", text):
        part = clean_display_text(line)
        if part:
            parts.append(part)
    return parts


def split_teacher_names(text: str) -> list[str]:
    names = []
    for part in re.split(r"[+＋]+", clean_display_text(text)):
        name = clean_person_name(part, teacher=True)
        if name and not is_generic_text(name):
            names.append(name)
    return names


def clean_person_name(text: str, teacher: bool = False) -> str:
    text = clean_display_text(text)
    text = re.sub(r"^[\d\s.\-–—_()]+", "", text)
    text = re.sub(r"^[#*•]+", "", text).strip()
    if ":" in text:
        left, right = text.split(":", 1)
        if is_generic_text(left) or contains_teacher_keyword(left):
            text = right
    if "：" in text:
        left, right = text.split("：", 1)
        if is_generic_text(left) or contains_teacher_keyword(left):
            text = right
    if teacher:
        text = re.sub(
            r"^(?:الأستاذة?|الاستاذة?|أستاذة?|استاذة?|المعلمة?|معلمة?|الشيخة?|شيخة?|الشيخ|شيخ|المدرسة?|مدرسة?|أ\.?|ا\.?)\s*[:：/\-–—]*\s*",
            "",
            text,
            flags=re.IGNORECASE,
        )
    text = clean_display_text(text)
    text = re.sub(r"^[\d\s.\-–—_()]+", "", text)
    text = re.sub(r"\s*[\-–—]\s*$", "", text)
    return clean_display_text(text)


def clean_halaqa_name(text: str) -> str:
    text = clean_display_text(text)
    if ":" in text:
        left, right = text.split(":", 1)
        if "حل" in normalize_for_match(left) or is_generic_text(left):
            text = right
    text = clean_display_text(text)
    if not contains_halaqa_keyword(text):
        return text
    return text


def derive_teacher_from_halaqa_name(halaqa_name: str) -> list[str]:
    text = normalize_for_match(halaqa_name)
    if "حلقه" not in text:
        return []
    display = clean_display_text(halaqa_name)
    display = re.sub(r"^(?:ال)?حلقة\s*", "", display)
    display = re.sub(r"^(?:ال)?حلقه\s*", "", display)
    return split_teacher_names(display)


def parse_grade_cell(text: str) -> list[int]:
    display = clean_display_text(text)
    normalized = normalize_for_match(display)
    if not normalized:
        return []
    if "الفئه" in normalized:
        return []

    grades = set()
    numeric_matches = re.findall(r"\d+", normalized)
    if numeric_matches:
        for match in numeric_matches:
            number = int(match)
            if 2 <= number <= 12:
                grades.add(number)

    for word, grade in GRADE_WORDS.items():
        if normalize_for_match(word) in normalized:
            grades.add(grade)

    if not grades:
        return []

    normalized_without_grade_words = f" {normalized} "
    for word in sorted(GRADE_WORDS, key=len, reverse=True):
        normalized_word = normalize_for_match(word)
        normalized_without_grade_words = normalized_without_grade_words.replace(f" {normalized_word} ", " ")
        normalized_without_grade_words = normalized_without_grade_words.replace(normalized_word, " ")
    for token in (
        "الصف",
        "صف",
        "grade",
        "grades",
        "class",
        "وال",
        "و",
        "الى",
        "ل",
        "من",
        "بكلوريا",
        "بكالوريا",
        "البكالوريا",
    ):
        normalized_without_grade_words = normalized_without_grade_words.replace(token, " ")
    normalized_without_grade_words = re.sub(r"[\d+\-/،,.\s]+", " ", normalized_without_grade_words).strip()

    # Person names sometimes contain words like "ثاني"; require the whole cell
    # to be grade-like instead of accepting grade words embedded in names.
    if normalized_without_grade_words:
        return []

    return sorted(grades)


def category_code_for_grades(grades: list[int]) -> str | None:
    if not grades:
        return None
    category_codes = {GRADE_TO_CATEGORY.get(grade) for grade in grades}
    category_codes.discard(None)
    if len(category_codes) == 1:
        return category_codes.pop()
    return None


def category_from_text(text: str) -> str | None:
    normalized = normalize_for_match(text)
    if "الفئه" not in normalized and "صف" not in normalized and "بكالوريا" not in normalized and "بكلوريا" not in normalized:
        return None

    named_categories = {
        "الاولي": "1",
        "اولى": "1",
        "الاولى": "1",
        "الثانيه": "2",
        "ثانيه": "2",
        "الثالثه": "3",
        "ثالثه": "3",
        "الرابعه": "4",
        "رابعه": "4",
        "الخامسه": "5",
        "خامسه": "5",
    }
    for word, code in named_categories.items():
        if word in normalized:
            return code

    grades = parse_grade_cell(text)
    return category_code_for_grades(grades)


def format_grade_label(grades: list[int], fallback: str = "") -> str:
    if grades:
        return "+".join(str(grade) for grade in grades)
    return clean_display_text(fallback)


def is_probable_name(text: str) -> bool:
    text = clean_person_name(text)
    normalized = normalize_for_match(text)
    if not normalized or len(normalized) < 2:
        return False
    if normalized.isdigit():
        return False
    if is_generic_text(text):
        return False
    if parse_grade_cell(text):
        return False
    blocked = ("الفئه", "الصف", "الحلقه", "طلاب", "طالبات", "العدد", "ملاحظ")
    if any(word in normalized for word in blocked):
        return False
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", text))


def unique_preserving_order(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        key = normalize_for_match(value)
        if key and key not in seen:
            seen.add(key)
            result.append(value)
    return result


def truncate_with_suffix(base: str, suffix: str, max_length: int) -> str:
    base = clean_display_text(base)
    if len(base) + len(suffix) <= max_length:
        return f"{base}{suffix}"
    return f"{base[: max_length - len(suffix)].rstrip()}{suffix}"


class SheetParser:
    def __init__(self, grid: list[list[str]], allow_empty_blocks: bool = False):
        self.grid = grid
        self.max_col = max((len(row) for row in grid), default=0)
        self.allow_empty_blocks = allow_empty_blocks
        self.issues: list[Issue] = []
        self.candidate_count = 0
        self.rejected_count = 0

    def parse(self) -> list[ImportBlock]:
        starts = self._find_block_starts()
        self.candidate_count = len(starts)
        blocks = self._build_blocks(starts)
        valid_start_keys = {block.start_key for block in blocks if self._block_is_importable(block)}
        self.rejected_count = len(starts) - len(valid_start_keys)

        # Rebuild after dropping weak candidates so rejected rows do not shorten
        # the student range of the valid block above them.
        if len(valid_start_keys) != len(starts):
            starts = [start for start in starts if self._start_key(start) in valid_start_keys]
            blocks = self._build_blocks(starts)

        importable_blocks = []
        for block in blocks:
            if not block.category_code:
                self.rejected_count += 1
                self.issues.append(
                    Issue("skipped_unclear_grade", block.row_number, block.grade_label, "تم تجاهل الصف لأن الصف لا يطابق فئة واضحة.")
                )
                continue
            if not block.teacher_names:
                self.rejected_count += 1
                self.issues.append(
                    Issue("skipped_missing_teacher", block.row_number, block.raw_header, "تم تجاهل الصف لأنه لا يحتوي اسم معلم واضحا.")
                )
                continue
            if not block.students and not self.allow_empty_blocks:
                self.rejected_count += 1
                self.issues.append(
                    Issue("skipped_no_students", block.row_number, block.halaqa_name, "تم تجاهل الحلقة لأنه لا يوجد طلاب تحتها.")
                )
                continue
            importable_blocks.append(block)
        return importable_blocks

    def _block_is_importable(self, block: ImportBlock) -> bool:
        if not block.category_code or not block.teacher_names:
            return False
        if not block.students and not self.allow_empty_blocks:
            return False
        return True

    def _find_block_starts(self) -> list[BlockStart]:
        starts = []
        context_category = None
        for row_index, row in enumerate(self.grid):
            if self._row_is_empty(row):
                continue

            joined = " ".join(cell for cell in row if cell)
            context_match = category_from_text(joined)
            if context_match:
                context_category = context_match

            grade_candidates = []
            for col_index, text in enumerate(row):
                grades = parse_grade_cell(text)
                if not grades:
                    continue
                category_code = category_code_for_grades(grades)
                grade_candidates.append((col_index, grades, category_code))

            for col_index, grades, category_code in grade_candidates:
                if not category_code:
                    continue
                if not self._has_valid_teacher_evidence(row, col_index):
                    continue
                if not self.allow_empty_blocks and not self._next_row_has_probable_names(row_index):
                    continue
                if self._score_header_row(row_index, col_index, grades) < 3:
                    continue
                starts.append(
                    BlockStart(
                        row_index=row_index,
                        anchor_col=col_index,
                        grade_col=col_index,
                        grades=grades,
                        category_code=category_code,
                        category_source="grade",
                    )
                )

            if grade_candidates:
                continue

        return starts

    def _score_header_row(self, row_index: int, grade_col: int, grades: list[int]) -> int:
        row = self.grid[row_index]
        texts = self._teacher_evidence_cells(row, grade_col)
        if not texts:
            return -10

        joined = " ".join(texts)
        score = 0
        if contains_teacher_keyword(joined):
            score += 3
        if "+" in joined or "＋" in joined:
            score += 2
        if contains_halaqa_keyword(joined):
            score += 2
        if len(grades) > 1:
            score += 1
        previous_is_boundary = self._previous_row_is_block_boundary(row_index)
        next_has_names = self._next_row_has_probable_names(row_index)
        if previous_is_boundary:
            score += 1
        if next_has_names:
            score += 1
        if previous_is_boundary and next_has_names and self._has_plain_teacher_name(texts):
            score += 1

        grade_text = row[grade_col]
        if re.fullmatch(r"\d+", normalize_for_match(grade_text)) and not (
            contains_teacher_keyword(joined) or contains_halaqa_keyword(joined) or "+" in joined
        ):
            if not previous_is_boundary:
                score -= 2

        return score

    def _previous_row_is_block_boundary(self, row_index: int) -> bool:
        if row_index == 0:
            return True
        previous_row = self.grid[row_index - 1]
        if self._row_is_empty(previous_row):
            return True
        previous_text = " ".join(previous_row)
        if category_from_text(previous_text):
            return True
        normalized = normalize_for_match(previous_text)
        return "المعلم" in normalized and "الصف" in normalized

    def _has_plain_teacher_name(self, texts: list[str]) -> bool:
        if len(texts) != 1:
            return False
        text = clean_person_name(texts[0], teacher=True)
        if not is_probable_name(text):
            return False
        normalized = normalize_for_match(text)
        return "طالب" not in normalized and "طالبه" not in normalized

    def _has_valid_teacher_evidence(self, row: list[str], grade_col: int) -> bool:
        cells = self._teacher_evidence_cells(row, grade_col)
        if not cells:
            return False
        teacher_names, halaqa_name = self._extract_header_names(cells)
        if teacher_names:
            return True
        if halaqa_name and derive_teacher_from_halaqa_name(halaqa_name):
            return True
        return False

    def _has_strong_teacher_evidence(self, row: list[str], grade_col: int) -> bool:
        cells = self._teacher_evidence_cells(row, grade_col)
        if not cells:
            return False
        joined = " ".join(cells)
        if contains_teacher_keyword(joined) or "+" in joined or "＋" in joined:
            return True
        teacher_names, halaqa_name = self._extract_header_names(cells)
        return bool(halaqa_name and teacher_names)

    def _teacher_evidence_cells(self, row: list[str], grade_col: int) -> list[str]:
        cells = []
        for col, cell in enumerate(row):
            if col == grade_col or not cell:
                continue
            if is_generic_text(cell) or parse_grade_cell(cell):
                continue
            if abs(col - grade_col) > 5 and not contains_teacher_keyword(cell) and "+" not in cell and not contains_halaqa_keyword(cell):
                continue
            cells.append(cell)
        return cells

    def _build_blocks(self, starts: list[BlockStart]) -> list[ImportBlock]:
        if not starts:
            return []

        starts_by_row = defaultdict(list)
        for start in starts:
            starts_by_row[start.row_index].append(start)

        spans = {}
        for row_index, row_starts in starts_by_row.items():
            row_starts.sort(key=lambda item: item.anchor_col)
            if len(row_starts) == 1:
                spans[self._start_key(row_starts[0])] = (0, self.max_col - 1)
                continue
            for index, start in enumerate(row_starts):
                previous_anchor = row_starts[index - 1].anchor_col if index else None
                next_anchor = row_starts[index + 1].anchor_col if index + 1 < len(row_starts) else None
                left = 0 if previous_anchor is None else ((previous_anchor + start.anchor_col) // 2) + 1
                right = self.max_col - 1 if next_anchor is None else ((start.anchor_col + next_anchor) // 2)
                spans[self._start_key(start)] = (left, right)

        all_start_rows = sorted({start.row_index for start in starts})
        blocks = []
        occurrence_counts = Counter()
        for start in starts:
            left, right = spans[self._start_key(start)]
            end_row = self._end_row_for_start(start.row_index, all_start_rows)
            block = self._build_block(start, left, right, end_row)
            occurrence_counts[normalize_for_match(block.halaqa_name)] += 1
            occurrence = occurrence_counts[normalize_for_match(block.halaqa_name)]
            if occurrence > 1:
                block.halaqa_name = truncate_with_suffix(block.halaqa_name, f" - {occurrence}", Halaqa._meta.get_field("name").max_length)
            blocks.append(block)
        return blocks

    def _start_key(self, start: BlockStart) -> tuple[int, int, int | None]:
        return (start.row_index, start.anchor_col, start.grade_col)

    def _build_block(self, start: BlockStart, left: int, right: int, end_row: int) -> ImportBlock:
        header_cells = [cell for cell in self.grid[start.row_index][left : right + 1] if cell]
        header_text = " | ".join(header_cells)
        teacher_names, halaqa_name = self._extract_header_names(header_cells)
        grade_label = format_grade_label(start.grades)
        if not halaqa_name:
            if teacher_names:
                halaqa_name = f"حلقة {' و '.join(teacher_names[:3])}"
            else:
                halaqa_name = f"حلقة صف {start.row_index + 1}"

        students = self._collect_students(start.row_index + 1, end_row, left, right)
        return ImportBlock(
            row_number=start.row_index + 1,
            left_col=left + 1,
            right_col=right + 1,
            start_key=self._start_key(start),
            grades=start.grades,
            grade_label=grade_label,
            category_code=start.category_code,
            teacher_names=teacher_names,
            halaqa_name=truncate_with_suffix(halaqa_name, "", Halaqa._meta.get_field("name").max_length),
            students=students,
            raw_header=header_text,
        )

    def _extract_header_names(self, header_cells: list[str]) -> tuple[list[str], str]:
        teacher_names = []
        halaqa_name = ""
        for cell in header_cells:
            if parse_grade_cell(cell) or is_generic_text(cell):
                continue
            if contains_halaqa_keyword(cell):
                possible_halaqa = clean_halaqa_name(cell)
                if possible_halaqa and not halaqa_name:
                    halaqa_name = possible_halaqa
                continue
            teacher_names.extend(split_teacher_names(cell))

        teacher_names = unique_preserving_order(teacher_names)
        if not teacher_names and halaqa_name:
            teacher_names = derive_teacher_from_halaqa_name(halaqa_name)
        return teacher_names, halaqa_name

    def _collect_students(self, start_row: int, end_row: int, left: int, right: int) -> list[str]:
        students = []
        for row_index in range(start_row, end_row + 1):
            row = self.grid[row_index][left : right + 1]
            if self._row_is_empty(row):
                break
            if self._looks_like_header_boundary(row):
                break
            for cell in row:
                for part in split_cell_parts(cell):
                    name = clean_person_name(part)
                    if is_probable_name(name):
                        students.append(name)
        return unique_preserving_order(students)

    def _looks_like_header_boundary(self, row: list[str]) -> bool:
        grade_cols = [index for index, cell in enumerate(row) if parse_grade_cell(cell)]
        if not grade_cols:
            return False
        for grade_col in grade_cols:
            if self._has_strong_teacher_evidence(row, grade_col):
                return True
        return False

    def _end_row_for_start(self, row_index: int, start_rows: list[int]) -> int:
        for start_row in start_rows:
            if start_row > row_index:
                return start_row - 1
        return len(self.grid) - 1

    def _next_row_has_probable_names(self, row_index: int) -> bool:
        for next_index in range(row_index + 1, min(row_index + 4, len(self.grid))):
            row = self.grid[next_index]
            if self._row_is_empty(row):
                continue
            return any(is_probable_name(cell) for cell in row)
        return False

    def _row_is_empty(self, row: list[str]) -> bool:
        return not any(clean_display_text(cell) for cell in row)

    def _first_non_empty_col(self, row: list[str]) -> int | None:
        for index, cell in enumerate(row):
            if cell:
                return index
        return None


class Command(BaseCommand):
    help = "Import summer 2026 halaqa distribution from an Excel Sheet1 file."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=str(Path(settings.BASE_DIR) / "import_data" / "توزيع صيف 2026.xlsx"),
            help="Excel file path. Defaults to backend/import_data/توزيع صيف 2026.xlsx",
        )
        parser.add_argument("--sheet", default="Sheet1", help="Worksheet name to import. Defaults to Sheet1.")

        mode = parser.add_mutually_exclusive_group()
        mode.add_argument("--dry-run", action="store_true", help="Parse and report without writing to the database.")
        mode.add_argument("--commit", action="store_true", help="Write changes to the database inside one transaction.")

        parser.add_argument("--create-missing-teachers", action="store_true")
        parser.add_argument("--create-missing-students", action="store_true")
        parser.add_argument("--default-birth-date", default="2010-01-01")
        parser.add_argument("--update-categories", action="store_true")
        parser.add_argument("--replace-teacher-assignments", action="store_true")
        parser.add_argument(
            "--allow-empty-halaqas",
            action="store_true",
            help="Import valid teacher/grade blocks even when no students are found underneath.",
        )
        parser.add_argument(
            "--debug-parsed-blocks",
            action="store_true",
            help="Print the accepted parsed halaqa blocks before importing.",
        )
        parser.add_argument("--conflicts-csv", default="")

    def handle(self, *args, **options):
        self.dry_run = not options["commit"]
        self.options = options
        self.stats = Counter()
        self.issues: list[Issue] = []
        self.issue_keys = set()
        self.planned_usernames = set()
        self.planned_teachers = {}
        self.planned_students = {}
        self.category_cache = {}
        self.parser_candidate_count = 0
        self.parser_rejected_count = 0

        try:
            self.default_birth_date = date.fromisoformat(options["default_birth_date"])
        except ValueError as exc:
            raise CommandError("--default-birth-date must be in YYYY-MM-DD format.") from exc

        file_path = self._resolve_file_path(options["file"])
        blocks = self._read_blocks(file_path, options["sheet"])
        self.stdout.write(
            self.style.NOTICE(
                f"تم اعتماد {len(blocks)} حلقة من أصل {self.parser_candidate_count} مرشح في الورقة {options['sheet']}."
            )
        )
        if self.parser_rejected_count:
            self.stdout.write(self.style.WARNING(f"تم تجاهل {self.parser_rejected_count} مرشح لا يملك دليلا كافيا."))

        for issue in getattr(self, "parser_issues", []):
            self._add_issue(issue.kind, issue.row, issue.value, issue.message)

        valid_blocks = [
            block
            for block in blocks
            if block.category_code and block.teacher_names and (block.students or options["allow_empty_halaqas"])
        ]
        if self.dry_run or options["debug_parsed_blocks"]:
            self._print_parsed_blocks(valid_blocks)

        if self.dry_run:
            self.stdout.write(self.style.WARNING("وضع التجربة فقط: لن يتم تعديل قاعدة البيانات."))
            self._process_blocks(valid_blocks)
        else:
            self.stdout.write(self.style.WARNING("وضع التنفيذ: سيتم الحفظ ضمن معاملة واحدة."))
            try:
                with transaction.atomic():
                    self._process_blocks(valid_blocks)
            except ImportRollback:
                raise CommandError("تم إلغاء الاستيراد.")

        self._write_conflicts_csv(options.get("conflicts_csv", ""))
        self._print_summary()

    def _resolve_file_path(self, requested_path: str) -> Path:
        path = Path(requested_path)
        if not path.is_absolute():
            cwd_candidate = Path.cwd() / path
            base_candidate = Path(settings.BASE_DIR) / path
            project_candidate = Path(settings.BASE_DIR).parent / path
            if cwd_candidate.exists():
                path = cwd_candidate
            elif base_candidate.exists():
                path = base_candidate
            elif project_candidate.exists():
                path = project_candidate
            else:
                path = base_candidate
        if not path.exists():
            raise CommandError(f"ملف Excel غير موجود: {path}")
        return path

    def _read_blocks(self, file_path: Path, sheet_name: str) -> list[ImportBlock]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("حزمة openpyxl غير مثبتة. شغل: pip install -r requirements.txt") from exc

        workbook = load_workbook(file_path, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f"الورقة غير موجودة: {sheet_name}. الأوراق المتاحة: {', '.join(workbook.sheetnames)}")

        worksheet = workbook[sheet_name]
        grid = [
            [cell_to_text(worksheet.cell(row=row, column=col).value) for col in range(1, worksheet.max_column + 1)]
            for row in range(1, worksheet.max_row + 1)
        ]

        for merged_range in worksheet.merged_cells.ranges:
            value = cell_to_text(worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value)
            if not value:
                continue
            for row in range(merged_range.min_row - 1, merged_range.max_row):
                for col in range(merged_range.min_col - 1, merged_range.max_col):
                    if row < len(grid) and col < len(grid[row]) and not grid[row][col]:
                        grid[row][col] = value

        grid = self._trim_grid(grid)
        parser = SheetParser(grid, allow_empty_blocks=self.options["allow_empty_halaqas"])
        blocks = parser.parse()
        self.parser_candidate_count = parser.candidate_count
        self.parser_rejected_count = parser.rejected_count
        self.parser_issues = parser.issues
        return blocks

    def _trim_grid(self, grid: list[list[str]]) -> list[list[str]]:
        while grid and not any(grid[-1]):
            grid.pop()
        max_col = 0
        for row in grid:
            for index, cell in enumerate(row):
                if cell:
                    max_col = max(max_col, index + 1)
        return [row[:max_col] for row in grid]

    def _process_blocks(self, blocks: list[ImportBlock]):
        self.teacher_index = self._build_teacher_index()
        self.student_index = self._build_student_index()
        self._ensure_categories({block.category_code for block in blocks if block.category_code})

        for block in blocks:
            self._process_block(block)

    def _build_teacher_index(self):
        index = defaultdict(list)
        for teacher in Teacher.objects.select_related("user").all():
            index[normalize_for_match(teacher.full_name)].append(teacher)
        return index

    def _build_student_index(self):
        index = defaultdict(list)
        for student in Student.objects.select_related("category", "halaqa").all():
            index[normalize_for_match(student.name)].append(student)
        return index

    def _ensure_categories(self, category_codes: set[str]):
        for code in sorted(category_codes):
            spec = CATEGORY_SPECS[code]
            category = Category.objects.filter(code=code).first()
            if category is None:
                self.stats["categories_create"] += 1
                if not self.dry_run:
                    category = Category.objects.create(
                        code=code,
                        name=spec["name"],
                        grade_span=spec["grade_span"],
                        display_order=spec["display_order"],
                        is_special=spec["is_special"],
                        notes="",
                    )
                else:
                    category = VirtualRecord(name=spec["name"], category_id=None)
                self.category_cache[code] = category
                continue

            changed = (
                category.name != spec["name"]
                or category.grade_span != spec["grade_span"]
                or category.display_order != spec["display_order"]
                or category.is_special != spec["is_special"]
            )
            if changed and self.options["update_categories"]:
                self.stats["categories_update"] += 1
                if not self.dry_run:
                    category.name = spec["name"]
                    category.grade_span = spec["grade_span"]
                    category.display_order = spec["display_order"]
                    category.is_special = spec["is_special"]
                    category.save(update_fields=["name", "grade_span", "display_order", "is_special"])
            elif changed:
                self.stats["categories_reuse_without_update"] += 1
                self._add_issue(
                    "category_diff",
                    None,
                    code,
                    "الفئة موجودة لكن الاسم/النطاق مختلف. استخدم --update-categories لتحديثها.",
                )
            else:
                self.stats["categories_reuse"] += 1
            self.category_cache[code] = category

    def _process_block(self, block: ImportBlock):
        category = self._get_category(block.category_code)
        if category is None:
            self._add_issue("missing_category", block.row_number, block.category_code or "", "لم يتم العثور على الفئة.")
            return

        halaqa = self._resolve_halaqa(block, category)
        teachers = []
        for teacher_name in block.teacher_names:
            teacher = self._resolve_teacher(teacher_name, block.row_number)
            if teacher:
                teachers.append(teacher)

        if not teachers:
            self._add_issue("block_skipped", block.row_number, block.halaqa_name, "تم تخطي الحلقة لعدم وجود معلم صالح.")
            return

        for teacher in teachers:
            self._link_teacher_to_halaqa(teacher, halaqa, block.row_number)

        if not block.students:
            self._add_issue("no_students", block.row_number, block.halaqa_name, "لم يتم العثور على طلاب تحت هذه الحلقة.")

        for student_name in block.students:
            student = self._resolve_student(student_name, block, category, halaqa)
            if student:
                self._link_student_to_halaqa(student, block, category, halaqa)

    def _get_category(self, code: str | None):
        if not code:
            return None
        if code in self.category_cache:
            return self.category_cache[code]
        return Category.objects.filter(code=code).first()

    def _resolve_halaqa(self, block: ImportBlock, category: Category):
        name = truncate_with_suffix(block.halaqa_name, "", Halaqa._meta.get_field("name").max_length)
        halaqa = Halaqa.objects.filter(name=name).first()
        if halaqa is None:
            self.stats["halaqas_create"] += 1
            if self.dry_run:
                return VirtualRecord(name=name, category_id=category.id)
            halaqa = Halaqa.objects.create(name=name, category=category, is_active=True)
        else:
            self.stats["halaqas_reuse"] += 1
            if halaqa.category_id != category.id:
                self.stats["halaqas_update_category"] += 1
                if not self.dry_run:
                    halaqa.category = category
                    halaqa.save(update_fields=["category"])
        return halaqa

    def _resolve_teacher(self, teacher_name: str, row_number: int):
        normalized = normalize_for_match(teacher_name)
        matches = self.teacher_index.get(normalized, [])
        if len(matches) == 1:
            self.stats["teachers_reuse"] += 1
            return matches[0]
        if len(matches) > 1:
            self.stats["teachers_conflict"] += 1
            self._add_issue("duplicate_teacher", row_number, teacher_name, "يوجد أكثر من معلم بنفس الاسم بعد التطبيع.")
            return None

        if normalized in self.planned_teachers:
            self.stats["teachers_reuse_planned"] += 1
            return self.planned_teachers[normalized]

        if not self.options["create_missing_teachers"]:
            self.stats["teachers_missing"] += 1
            self._add_issue("missing_teacher", row_number, teacher_name, "المعلم غير موجود. أضف --create-missing-teachers لإنشائه.")
            return None

        self.stats["teachers_create"] += 1
        self.stats["teacher_default_values"] += 1
        if self.dry_run:
            teacher = VirtualRecord(name=teacher_name, full_name=teacher_name)
            self.planned_teachers[normalized] = teacher
            return teacher

        username = self._next_username("teacher")
        User = get_user_model()
        user = User(username=username, first_name=teacher_name[:150], email="")
        user.set_unusable_password()
        user.save()
        profile, _ = Profile.objects.get_or_create(user=user)
        if profile.role != "teacher":
            profile.role = "teacher"
            profile.save(update_fields=["role"])
        teacher = Teacher.objects.create(user=user, full_name=teacher_name, phone="", qualification="")
        self.teacher_index[normalized].append(teacher)
        self.planned_teachers[normalized] = teacher
        return teacher

    def _resolve_student(self, student_name: str, block: ImportBlock, category: Category, halaqa):
        normalized = normalize_for_match(student_name)
        matches = self.student_index.get(normalized, [])
        if len(matches) == 1:
            self.stats["students_reuse"] += 1
            return matches[0]
        if len(matches) > 1:
            self.stats["students_conflict"] += 1
            self._add_issue("duplicate_student", block.row_number, student_name, "يوجد أكثر من طالب بنفس الاسم بعد التطبيع، لذلك تم تخطيه.")
            return None

        if normalized in self.planned_students:
            self.stats["students_reuse_planned"] += 1
            return self.planned_students[normalized]

        if not self.options["create_missing_students"]:
            self.stats["students_missing"] += 1
            self._add_issue("missing_student", block.row_number, student_name, "الطالب غير موجود. أضف --create-missing-students لإنشائه.")
            return None

        self.stats["students_create"] += 1
        self.stats["student_default_birth_date"] += 1
        if self.dry_run:
            student = VirtualRecord(name=student_name)
            self.planned_students[normalized] = student
            return student

        student = Student.objects.create(
            name=student_name,
            birth_date=self.default_birth_date,
            parent=None,
            parent_phone="",
            address="",
            grade=block.grade_label,
            category=category,
            halaqa=halaqa,
            enrollment_date=timezone.localdate(),
            previous_memorization_amount=0,
        )
        self.student_index[normalized].append(student)
        self.planned_students[normalized] = student
        return student

    def _link_teacher_to_halaqa(self, teacher, halaqa, row_number: int):
        if self.dry_run:
            self.stats["teacher_halaqa_links"] += 1
            return

        through_model = Halaqa.teachers.through
        _, created = through_model.objects.get_or_create(teacher_id=teacher.pk, halaqa_id=halaqa.pk)
        if created:
            self.stats["teacher_halaqa_links"] += 1
        else:
            self.stats["teacher_halaqa_links_reuse"] += 1

        active_assignment = TeacherAssignment.objects.filter(teacher=teacher, is_active=True).select_related("halaqa").first()
        if active_assignment is None:
            TeacherAssignment.objects.create(teacher=teacher, halaqa=halaqa, start_date=timezone.localdate(), is_active=True)
            Teacher.objects.filter(pk=teacher.pk, current_halaqa__isnull=True).update(current_halaqa=halaqa)
            self.stats["teacher_assignments_create"] += 1
            return

        if active_assignment.halaqa_id == halaqa.pk:
            self.stats["teacher_assignments_reuse"] += 1
            return

        if self.options["replace_teacher_assignments"]:
            active_assignment.is_active = False
            active_assignment.end_date = timezone.localdate()
            active_assignment.save(update_fields=["is_active", "end_date"])
            TeacherAssignment.objects.create(teacher=teacher, halaqa=halaqa, start_date=timezone.localdate(), is_active=True)
            Teacher.objects.filter(pk=teacher.pk).update(current_halaqa=halaqa)
            self.stats["teacher_assignments_replaced"] += 1
            return

        self.stats["teacher_assignment_conflicts"] += 1
        self._add_issue(
            "teacher_assignment_conflict",
            row_number,
            teacher.full_name,
            f"لدى المعلم إسناد نشط في حلقة أخرى: {active_assignment.halaqa.name}. تم ربطه ManyToMany فقط.",
        )

    def _link_student_to_halaqa(self, student, block: ImportBlock, category: Category, halaqa):
        if isinstance(student, VirtualRecord):
            self.stats["student_halaqa_links"] += 1
            return

        halaqa_pk = getattr(halaqa, "pk", None)
        active_membership = (
            HalaqaMembership.objects.filter(student=student, is_active=True)
            .select_related("halaqa")
            .order_by("-join_date", "-id")
            .first()
        )
        if active_membership and active_membership.halaqa_id != halaqa_pk:
            self.stats["student_membership_conflicts"] += 1
            self._add_issue(
                "student_membership_conflict",
                block.row_number,
                student.name,
                f"لدى الطالب عضوية نشطة في حلقة أخرى: {active_membership.halaqa.name}. تم تخطي نقله.",
            )
            return

        if self.dry_run:
            self.stats["student_halaqa_links"] += 1
            return

        changed_fields = []
        if student.grade != block.grade_label:
            student.grade = block.grade_label
            changed_fields.append("grade")
        if student.category_id != category.id:
            student.category = category
            changed_fields.append("category")
        if student.halaqa_id != halaqa.id:
            student.halaqa = halaqa
            changed_fields.append("halaqa")
        if changed_fields:
            student.save(update_fields=changed_fields)
            self.stats["students_update"] += 1

        same_membership = HalaqaMembership.objects.filter(student=student, halaqa=halaqa).first()
        if same_membership:
            if not same_membership.is_active:
                same_membership.is_active = True
                same_membership.end_date = None
                same_membership.save(update_fields=["is_active", "end_date"])
                self.stats["student_memberships_reactivate"] += 1
            else:
                self.stats["student_memberships_reuse"] += 1
            return

        HalaqaMembership.objects.create(student=student, halaqa=halaqa, is_active=True)
        self.stats["student_memberships_create"] += 1

    def _next_username(self, prefix: str) -> str:
        User = get_user_model()
        counter = 1
        while True:
            username = f"{prefix}_{counter:03d}"
            if username not in self.planned_usernames and not User.objects.filter(username=username).exists():
                self.planned_usernames.add(username)
                return username
            counter += 1

    def _add_issue(self, kind: str, row: int | None, value: str, message: str):
        cleaned_value = clean_display_text(value)
        issue_key = (kind, row, normalize_for_match(cleaned_value), message)
        if issue_key in self.issue_keys:
            return
        self.issue_keys.add(issue_key)
        self.issues.append(Issue(kind=kind, row=row, value=cleaned_value, message=message))

    def _print_parsed_blocks(self, blocks: list[ImportBlock]):
        if not blocks:
            self.stdout.write(self.style.WARNING("لا توجد حلقات مقبولة بعد فحص الصف/المعلم/الطلاب."))
            return

        self.stdout.write("")
        self.stdout.write(self.style.NOTICE("الحلقات التي فهمها المستورد"))
        header = f"{'row':>5}  {'grade':<8}  {'category':<14}  {'students':>8}  {'halaqa':<32}  teachers"
        self.stdout.write(header)
        self.stdout.write("-" * len(header))
        for block in blocks:
            category_name = CATEGORY_SPECS.get(block.category_code, {}).get("name", block.category_code or "")
            halaqa_name = truncate_with_suffix(block.halaqa_name, "", 32)
            teacher_names = " + ".join(block.teacher_names)
            self.stdout.write(
                f"{block.row_number:>5}  {block.grade_label:<8}  {category_name:<14}  "
                f"{len(block.students):>8}  {halaqa_name:<32}  {teacher_names}"
            )

    def _write_conflicts_csv(self, path_value: str):
        if not path_value:
            return
        path = Path(path_value)
        if not path.is_absolute():
            path = Path(settings.BASE_DIR) / path
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["kind", "row", "value", "message"])
            writer.writeheader()
            for issue in self.issues:
                writer.writerow(
                    {
                        "kind": issue.kind,
                        "row": issue.row or "",
                        "value": issue.value,
                        "message": issue.message,
                    }
                )
        self.stdout.write(self.style.SUCCESS(f"تمت كتابة تقرير التعارضات: {path}"))

    def _print_summary(self):
        self.stdout.write("")
        self.stdout.write(self.style.NOTICE("ملخص الاستيراد"))
        ordered_keys = [
            ("categories_create", "فئات سيتم إنشاؤها/أنشئت"),
            ("categories_reuse", "فئات معاد استخدامها"),
            ("categories_update", "فئات سيتم تحديثها/حدثت"),
            ("categories_reuse_without_update", "فئات مختلفة أُعيد استخدامها بدون تحديث"),
            ("halaqas_create", "حلقات سيتم إنشاؤها/أنشئت"),
            ("halaqas_reuse", "حلقات معاد استخدامها"),
            ("halaqas_update_category", "حلقات ستتغير فئتها/تغيرت"),
            ("teachers_create", "معلمون سيتم إنشاؤهم/أنشئوا"),
            ("teachers_reuse", "معلمون معاد استخدامهم"),
            ("teachers_missing", "معلمون مفقودون"),
            ("teachers_conflict", "تعارض أسماء معلمين"),
            ("teacher_halaqa_links", "روابط معلمين بحلقات"),
            ("teacher_assignment_conflicts", "تعارض إسناد معلم نشط"),
            ("students_create", "طلاب سيتم إنشاؤهم/أنشئوا"),
            ("students_reuse", "طلاب معاد استخدامهم"),
            ("students_update", "طلاب حُدثوا"),
            ("students_missing", "طلاب مفقودون"),
            ("students_conflict", "تعارض أسماء طلاب"),
            ("student_halaqa_links", "روابط طلاب بحلقات"),
            ("student_membership_conflicts", "تعارض عضوية طالب نشطة"),
            ("student_default_birth_date", f"طلاب استخدم لهم تاريخ ميلاد افتراضي ({self.default_birth_date})"),
            ("teacher_default_values", "معلمون استخدمت لهم بيانات افتراضية"),
        ]
        for key, label in ordered_keys:
            self.stdout.write(f"- {label}: {self.stats.get(key, 0)}")

        if self.issues:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING(f"تنبيهات/تعارضات: {len(self.issues)}"))
            for issue in self.issues[:50]:
                row = f"صف {issue.row}: " if issue.row else ""
                self.stdout.write(f"- [{issue.kind}] {row}{issue.value} - {issue.message}")
            if len(self.issues) > 50:
                self.stdout.write(f"... وهناك {len(self.issues) - 50} تنبيها إضافيا. استخدم --conflicts-csv لحفظها كاملة.")
